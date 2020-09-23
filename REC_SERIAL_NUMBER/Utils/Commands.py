import requests
from Utils.Variables import *
import errno
import os
import logging
import logging.handlers
import urllib3
# from openpyxl import load_workbook
# from openpyxl import Workbook
# from openpyxl.styles import Font
# import time
import datetime
from datetime import datetime
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
# from openpyxl.styles import Color, Fill
# from openpyxl.cell import Cell
import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
from email.MIMEBase import MIMEBase
from email import encoders



urllib3.disable_warnings()
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

currday = datetime.now().today().date().strftime("%d_%m_%Y")
currmonth = datetime.now().today().date().strftime("%m_%Y")
currtime = str(datetime.now())
currtime =  currtime.replace(" ", "_").replace(":", "_").replace("-", "_").replace(".", "_")


if __name__ == "__main__":
    fileDir = os.path.dirname(os.path.dirname(os.path.realpath('__file__')))
else:
    fileDir = os.path.dirname(os.path.realpath('__file__'))

rec_book = fileDir + '/' + currmonth + '_Record.xlsx'
logfile_dir = fileDir + "/LOGS/"+ vd_dict['ip'] + "_" + currtime + "/"
if not os.path.exists(os.path.dirname(logfile_dir)):
    try:
        os.mkdir(os.path.dirname(logfile_dir))
    except OSError as exc:  # Guard against race condition
        if exc.errno != errno.EEXIST:
            raise
print fileDir
daily_report_book = logfile_dir + currday + "_Record.xlsx"
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
formatter1 = logging.Formatter("%(message)s")
console = logging.StreamHandler()
console.setLevel(logging.DEBUG)
console.setFormatter(formatter1)
logging.getLogger('').addHandler(console)


def setup_logger(name, filename, level=logging.DEBUG):
    log_file = logfile_dir + filename  + ".log"
    handler = logging.FileHandler(log_file)
    handler.setFormatter(formatter)
    logger = logging.getLogger(name)
    logger.setLevel(level)
    logger.addHandler(handler)
    logger = logging.getLogger(name)
    return logger

main_logger = setup_logger('Main', 'UpgradeVersaCpes')


def get_device_dict():
    global batch
    response1 = requests.get(vdurl + appliance_url,
                             auth=(user, passwd),
                             headers=headers3,
                             verify=False)
    data1 = response1.json()
    count, day, batch = 1, 1, 1
    #print data1
    device_dict = {}
    for i in data1['versanms.ApplianceStatusResult']['appliances']:
        #print i
	#raw_input()
        try:
            if i['Hardware']!="":
                device_dict[i['name']] = i['Hardware']['serialNo']
        except KeyError as ke:
            print i['name']
            print "Hardware Info NIL"
            device_dict[i['name']] = "NIL"
        #print count, day, batch
        count +=1
    # print devices_list
    #main_logger.info(device_dict)
    return device_dict

def write_excel_sheet(data_dict):
    try:
        book = load_workbook(rec_book)
    except:
        book = Workbook()

    # currtime = str(datetime.now())
    # currtime =  currtime.replace(" ", "_").replace(":", "_").replace("-", "_").replace(".", "_")
    # print currtime
    # Day = currtime
    sheet = book.create_sheet(currday, 0)
    sheet.cell(column=1, row=1, value='DEVICE')
    sheet.cell(column=2, row=1, value='SERIAL_NO')
    i=2
    for k, v in data_dict.iteritems():
        sheet.cell(column=1, row=i, value=k)
        sheet.cell(column=2, row=i, value=v)
        i+=1
    book.save(rec_book)
    #book.close()
    try:
        daily_book = load_workbook(daily_report_book)
    except:
        daily_book = Workbook()
    sheet1 = daily_book.create_sheet(currday, 0)
    sheet1.cell(column=1, row=1, value='DEVICE')
    sheet1.cell(column=2, row=1, value='SERIAL_NO')
    i=2
    for k, v in data_dict.iteritems():
        sheet1.cell(column=1, row=i, value=k)
        sheet1.cell(column=2, row=i, value=v)
        i+=1
    daily_book.save(daily_report_book)
    #daily_book.close()

def send_mail():
    fromaddr = "nv-bh01-pgt@colt.net"
    #toaddr = ["sathishkumar.murugesan@colt.net"]
    toaddr = ["Stefano.Campostrini@colt.net", "sathishkumar.murugesan@colt.net", "Radu.Dragomir@colt.net", "Manish.Kumar@colt.net", \
		"Madhusudan.Narayanaswamy@colt.net", "Danny.Pinto@colt.net", "Ravidutt.Sharma@colt.net, Sultan.Shaikh@colt.net"]
    #emails = ["mike@somewhere.org", "nedry@jp.net"]

    msg = MIMEMultipart()

    msg['From'] = fromaddr
    msg['To'] = (', ').join(toaddr)
    msg['Subject'] = "Production VD appliance's Serial-number Recorded on " + currday

    body = "Hi Team,\n" \
           "Current day & month record is attached in this Mail. \n" \
           "ALL Records Logged & saved in  <bastion host : nv-bh01-pgt > <folder:/opt/script/REC_SERIAL_NUMBER/>.\n" \
           "<EOM>"

    msg.attach(MIMEText(body, 'plain'))

    # filename = "Record.xlsx"
    attachment = open(rec_book, "rb")

    part = MIMEBase('application', 'octet-stream')
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= %s" % currmonth + "_record.xlsx")

    msg.attach(part)

    server = smtplib.SMTP('nv-mo01-pgt.nv.colt.net', 25)
    server.starttls()
    text = msg.as_string()
    server.sendmail(fromaddr, toaddr, text)
    server.quit()


def Rec_ser_num():
    write_excel_sheet(get_device_dict())
    send_mail()


# main()



