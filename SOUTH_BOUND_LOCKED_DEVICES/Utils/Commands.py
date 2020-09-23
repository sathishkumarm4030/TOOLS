import requests
from Utils.Variables import *
import errno
import os
import logging
import logging.handlers
import urllib3
import time
from netmiko import redispatch
from netmiko import ConnectHandler
# from openpyxl import load_workbook
# from openpyxl import Workbook
# from openpyxl.styles import Font
# import time
import datetime
from datetime import datetime
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
# from openpyxl.styles import Color, Fill
# from openpyxl.cell import Cell
import smtplib
# from email.MIMEMultipart import MIMEMultipart
# from email.MIMEText import MIMEText
# from email.MIMEBase import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import re


urllib3.disable_warnings()
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

currday = datetime.now().today().date().strftime("%d_%m_%Y")
currmonth = datetime.now().today().date().strftime("%m_%Y")
currtime = str(datetime.now())
currtime =  currtime.replace(" ", "_").replace(":", "_").replace("-", "_").replace(".", "_")


if __name__ == "__main__":
    fileDir = os.path.dirname(os.path.dirname(os.path.realpath('__file__')))
else:
    fileDir = os.path.dirname(os.path.realpath('__file__'))

rec_book = fileDir + '/' + currmonth + '_southboundlocked.xlsx'
logfile_dir = fileDir + "/LOGS/"+ vd_dict['ip'] + "_" + currtime + "/"
if not os.path.exists(os.path.dirname(logfile_dir)):
    try:
        os.mkdir(os.path.dirname(logfile_dir))
    except OSError as exc:  # Guard against race condition
        if exc.errno != errno.EEXIST:
            raise
print fileDir
daily_report_book = logfile_dir + currday + "_southboundlocked.xlsx"
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
formatter1 = logging.Formatter("%(message)s")
console = logging.StreamHandler()
console.setLevel(logging.DEBUG)
console.setFormatter(formatter1)
logging.getLogger('').addHandler(console)


def setup_logger(name, filename, level=logging.DEBUG):
    global log_file
    log_file = logfile_dir + filename  + ".log"
    handler = logging.FileHandler(log_file)
    handler.setFormatter(formatter)
    logger = logging.getLogger(name)
    logger.setLevel(level)
    logger.addHandler(handler)
    logger = logging.getLogger(name)
    return logger

main_logger = setup_logger('Main', 'VD')

def make_connection(a_device):
    global main_logger
    try:
        net_connect = ConnectHandler(global_delay_factor=5, **a_device)
        output = net_connect.read_channel()
        main_logger.info(output)
    except ValueError as Va:
        main_logger.info(Va)
        main_logger.info("Not able to enter Versa Director CLI. please Check")
        exit()
    # net_connect.enable()
    time.sleep(5)
    main_logger.debug("{}: {}".format(net_connect.device_type, net_connect.find_prompt()))
    # print str(net_connect) + " connection opened"
    main_logger.debug(str(net_connect) + " connection opened")
    return net_connect


def get_device_dict():
    global batch
    response1 = requests.get(vdurl + appliance_url,
                             auth=(user, passwd),
                             headers=headers3,
                             verify=False)
    data1 = response1.json()
    count, day, batch = 1, 1, 1
    #print data1
    device_dict = {}
    for i in data1['versanms.ApplianceStatusResult']['appliances']:
        #print i
	#raw_input()
        try:
            if i['Hardware']!="":
                device_dict[i['name']] = i['Hardware']['serialNo']
        except KeyError as ke:
            print i['name']
            print "Hardware Info NIL"
            device_dict[i['name']] = "NIL"
        #print count, day, batch
        count +=1
    # print devices_list
    #main_logger.info(device_dict)
    return device_dict

def write_excel_sheet(data_dict):
    try:
        book = load_workbook(rec_book)
    except:
        book = Workbook()

    # currtime = str(datetime.now())
    # currtime =  currtime.replace(" ", "_").replace(":", "_").replace("-", "_").replace(".", "_")
    # print currtime
    # Day = currtime
    sheet = book.create_sheet(currday, 0)
    sheet.cell(column=1, row=1, value='DEVICE')
    sheet.cell(column=2, row=1, value='ADMIN_STATE')
    i=2
    for k, v in data_dict.iteritems():
        sheet.cell(column=1, row=i, value=k)
        sheet.cell(column=2, row=i, value=v)
        i+=1
    book.save(rec_book)
    #book.close()
    try:
        daily_book = load_workbook(daily_report_book)
    except:
        daily_book = Workbook()
    sheet1 = daily_book.create_sheet(currday, 0)
    sheet1.cell(column=1, row=1, value='DEVICE')
    sheet1.cell(column=2, row=1, value='ADMIN_STATE')
    i=2
    for k, v in data_dict.iteritems():
        sheet1.cell(column=1, row=i, value=k)
        sheet1.cell(column=2, row=i, value=v)
        i+=1
    daily_book.save(daily_report_book)
    #daily_book.close()

def send_mail():
    global output, log_file
    fromaddr = "nv-bh01-pgt@colt.net"
    #toaddr = ["sathishkumar.murugesan@colt.net"]
    toaddr = ["Stefano.Campostrini@colt.net", "sathishkumar.murugesan@colt.net", "Radu.Dragomir@colt.net", \
              "Manish.Kumar@colt.net", "Anand.Chaluvaiah@colt.net", "Mahesh.Gollapudi@colt.net"]
    #emails = ["mike@somewhere.org", "nedry@jp.net"]

    msg = MIMEMultipart()

    msg['From'] = fromaddr
    msg['To'] = (', ').join(toaddr)
    msg['Subject'] = "South Bound Locked Devices Recorded on " + currday

    body = "Hi Team,\n" \
            "Current day & month record is attached in this Mail. \n" \
           "ALL Records Logged & saved in  <bastion host : nv-bh01-pgt > <folder:/opt/script/SOUTH_BOUND_LOCKED_DEVICES/>.\n" \
           "<EOM>"
    msg.attach(MIMEText(body, 'plain'))

    # filename = "Record.xlsx"
    attachment = open(rec_book, "rb")

    part = MIMEBase('application', 'octet-stream')
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= %s" % currmonth + "_southboundlocked.xlsx")
    msg.attach(part)

    # filename = ".log"
    attachment1 = open(log_file, "rb")
    part1 = MIMEBase('application', 'octet-stream')
    part1.set_payload((attachment1).read())
    encoders.encode_base64(part1)
    part1.add_header('Content-Disposition', "attachment; filename= %s" % currmonth + "_VD.log")
    msg.attach(part1)

    server = smtplib.SMTP('nv-mo01-pgt.nv.colt.net', 25)
    #server = smtplib.SMTP('10.91.140.228', 25)
    server.starttls()
    text = msg.as_string()
    server.sendmail(fromaddr, toaddr, text)
    server.quit()


def Rec_ser_num():
    write_excel_sheet(get_device_dict())
    send_mail()

def Southbound_locked():
    global output
    netconnect = make_connection(vd_ssh_dict)
    cmd = "show configuration devices device * state admin-state | match southbound-locked | display set | except ScratchPad | except DataStore | nomore"
    output = netconnect.send_command_expect(cmd, max_loops=5000, strip_prompt=False, strip_command=False)
    main_logger.info(output)
    device_dict = {}
    #output_list = output.split("\n")
    # for i in output_list:
    #     if "state admin-state southbound-locked" in i:
    #         i = i.replace("set devices device ", "").replace("state admin-state southbound-locked", "")
    #         j = i.split()[0]
    #         print j
    #         device_dict[j] = "southbound-locked"
    devices = re.findall(r'set devices device (\S+) state admin-state southbound-locked', output)
    for device in devices:
        device_dict[device] = "southbound-locked"
    write_excel_sheet(device_dict)
    send_mail()
# main()



